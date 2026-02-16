from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, jsonify
from flask_login import login_required, current_user
from sqlalchemy import or_, func
from datetime import datetime, timedelta

from ..audit import log_audit
from .. import db
from ..utils import require_perm
from ..models import (
    Lead, LeadStatus, LeadSource, User,
    LeadActivity, ActivityType,
    Client, ClientBranch, EmployeeProfile, Industry, LeadService,
)

from werkzeug.utils import secure_filename


leads_bp = Blueprint("leads", __name__, template_folder="../templates")


# -------------------------
# Access helpers
# -------------------------
def _team_user_ids(manager_user_id: int):
    """Return direct + nested reportees user_ids (recursive)."""
    seen = set()
    stack = [manager_user_id]

    while stack:
        mid = stack.pop()
        rows = (EmployeeProfile.query
                .filter(EmployeeProfile.reporting_manager_user_id == mid)
                .with_entities(EmployeeProfile.user_id)
                .all())
        child_ids = [r[0] for r in rows if r[0] is not None and r[0] not in seen]
        for cid in child_ids:
            seen.add(cid)
            stack.append(cid)

    return list(seen)


def _allowed_lead_owner_ids():
    """
    Who's leads can the current user see?
    - If user has leads.view_all => everyone
    - Else => self + reportees
    """
    if current_user.has_perm("leads.view_all"):
        return None  # None means "no restriction"

    team_ids = _team_user_ids(current_user.id)
    return [current_user.id] + team_ids


def _owner_options_for_current_user():
    """Users to show in Owner dropdowns (list & assign)."""
    allowed = _allowed_lead_owner_ids()
    if allowed is None:
        return User.query.filter_by(is_active=True).order_by(User.name.asc()).all()
    return User.query.filter(User.id.in_(allowed)).order_by(User.name.asc()).all()


def _enforce_lead_access(lead: Lead):
    allowed = _allowed_lead_owner_ids()
    if allowed is None:
        return True
    if lead.owner_id in allowed:
        return True
    abort(403)


# -------------------------
# Helpers
# -------------------------
def _lead_code_next():
    last = db.session.query(Lead).order_by(Lead.id.desc()).first()
    nxt = (last.id + 1) if last else 1
    return f"LD-{nxt:06d}"


def _parse_date(v):
    if not v:
        return None
    return datetime.strptime(v, "%Y-%m-%d").date()


def _load_form_masters():
    """Keep masters in one place so we don't forget industries on error renders."""
    statuses = LeadStatus.query.filter_by(is_active=True).order_by(LeadStatus.sort_order.asc()).all()
    sources = LeadSource.query.filter_by(is_active=True).order_by(LeadSource.sort_order.asc()).all()
    clients = Client.query.filter_by(is_active=True).order_by(Client.company_name.asc()).all()
    industries = Industry.query.filter_by(is_active=True).order_by(Industry.sort_order.asc(), Industry.name.asc()).all()
    services = LeadService.query.filter_by(is_active=True).order_by(LeadService.sort_order.asc(), LeadService.name.asc()).all()
    return statuses, sources, clients, industries, services


# -------------------------
# Routes
# -------------------------
@leads_bp.route("/")
@login_required
@require_perm("leads.view")
def list_leads():
    q = (request.args.get("q") or "").strip()
    status_id = request.args.get("status_id") or ""
    source_id = request.args.get("source_id") or ""
    owner = request.args.get("owner") or ""
    service_id = request.args.get("service_id") or ""

    page = int(request.args.get("page", 1))
    per_page = 10

    query = Lead.query

    # Restrict visibility by default
    allowed_owner_ids = _allowed_lead_owner_ids()
    if allowed_owner_ids is not None:
        query = query.filter(Lead.owner_id.in_(allowed_owner_ids))

    # Search filters
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Lead.name.ilike(like),
            Lead.company.ilike(like),
            Lead.email.ilike(like),
            Lead.phone.ilike(like),
            Lead.lead_code.ilike(like),
        ))

    if status_id.isdigit():
        query = query.filter(Lead.status_id == int(status_id))

    if source_id.isdigit():
        query = query.filter(Lead.source_id == int(source_id))

    if service_id.isdigit():
        query = query.filter(Lead.service_id == int(service_id))

    # Owner filter
    if owner == "me":
        query = query.filter(Lead.owner_id == current_user.id)
    elif owner.isdigit():
        oid = int(owner)
        if allowed_owner_ids is None or oid in allowed_owner_ids:
            query = query.filter(Lead.owner_id == oid)
        else:
            query = query.filter(Lead.owner_id == current_user.id)

    query = query.order_by(Lead.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    services = LeadService.query.filter_by(is_active=True).order_by(LeadService.sort_order.asc(), LeadService.name.asc()).all()
    statuses = LeadStatus.query.filter_by(is_active=True).order_by(LeadStatus.sort_order.asc()).all()
    sources = LeadSource.query.filter_by(is_active=True).order_by(LeadSource.sort_order.asc()).all()
    owner_options = _owner_options_for_current_user()

    return render_template(
        "leads/list.html",
        pagination=pagination,
        q=q,
        status_id=status_id,
        source_id=source_id,
        service_id=service_id,
        owner=owner,
        statuses=statuses,
        sources=sources,
        services=services,
        owner_options=owner_options,
        allowed_owner_ids=allowed_owner_ids,
    )


@leads_bp.route("/new", methods=["GET", "POST"])
@login_required
@require_perm("leads.create")
def create_lead():
    statuses, sources, clients, industries, services = _load_form_masters()

    if request.method == "POST":
        client_id_raw = (request.form.get("client_id") or "").strip()
        branch_id_raw = (request.form.get("branch_id") or "").strip()
        client_id = int(client_id_raw) if client_id_raw.isdigit() else None
        branch_id = int(branch_id_raw) if branch_id_raw.isdigit() else None

        # Validate branch belongs to selected client
        if branch_id and not client_id:
            flash("Please select Client first before selecting Branch.", "danger")
            return render_template(
                "leads/form.html",
                mode="create", lead=None,
                statuses=statuses, sources=sources, clients=clients, industries=industries, services=services
            )

        if client_id and branch_id:
            br = ClientBranch.query.get(branch_id)
            if not br or br.client_id != client_id:
                flash("Selected branch does not belong to selected client.", "danger")
                return render_template(
                    "leads/form.html",
                    mode="create", lead=None,
                    statuses=statuses, sources=sources, clients=clients, industries=industries, services=services
                )

        lead = Lead(
            lead_code=_lead_code_next(),
            name=(request.form.get("name") or "").strip(),
            company=(request.form.get("company") or "").strip(),
            email=(request.form.get("email") or "").strip().lower(),
            phone_country=(request.form.get("phone_country") or "+91").strip(),
            phone=(request.form.get("phone") or "").strip(),
            location=(request.form.get("location") or "").strip(),
            industry_id=int(request.form.get("industry_id")) if (request.form.get("industry_id") or "").isdigit() else None,
            website=(request.form.get("website") or "").strip(),
            notes=(request.form.get("notes") or "").strip(),
            owner_id=current_user.id,
            status_id=int(request.form.get("status_id")) if (request.form.get("status_id") or "").isdigit() else None,
            source_id=int(request.form.get("source_id")) if (request.form.get("source_id") or "").isdigit() else None,
            service_id=int(request.form.get("service_id")) if (request.form.get("service_id") or "").isdigit() else None,
            client_id=client_id,
            branch_id=branch_id,
            estimated_closure_date=_parse_date(request.form.get("estimated_closure_date")),
        )

        if not lead.name:
            flash("Lead name is required.", "danger")
            return render_template(
                "leads/form.html",
                mode="create", lead=lead,
                statuses=statuses, sources=sources, clients=clients, industries=industries, services=services
            )

        db.session.add(lead)
        db.session.commit()

        log_audit("Lead", lead.id, "CREATE")
        flash("Lead created successfully ✅", "success")
        return redirect(url_for("leads.list_leads"))

    return render_template(
        "leads/form.html",
        mode="create", lead=None,
        statuses=statuses, sources=sources, clients=clients, industries=industries, services=services
    )


@leads_bp.route("/<int:lead_id>/edit", methods=["GET", "POST"])
@login_required
@require_perm("leads.edit")
def edit_lead(lead_id):
    lead = Lead.query.get_or_404(lead_id)
    _enforce_lead_access(lead)

    statuses, sources, clients, industries, services = _load_form_masters()

    if request.method == "POST":
        old_name = lead.name

        lead.name = (request.form.get("name") or "").strip()
        lead.company = (request.form.get("company") or "").strip()
        lead.email = (request.form.get("email") or "").strip().lower()
        lead.phone_country = (request.form.get("phone_country") or "+91").strip()
        lead.phone = (request.form.get("phone") or "").strip()
        lead.location = (request.form.get("location") or "").strip()
        lead.industry_id = int(request.form.get("industry_id")) if (request.form.get("industry_id") or "").isdigit() else None
        lead.website = (request.form.get("website") or "").strip()
        lead.notes = (request.form.get("notes") or "").strip()
        lead.service_id = int(request.form.get("service_id")) if (request.form.get("service_id") or "").isdigit() else None
        lead.status_id = int(request.form.get("status_id")) if (request.form.get("status_id") or "").isdigit() else None
        lead.source_id = int(request.form.get("source_id")) if (request.form.get("source_id") or "").isdigit() else None

        client_id_raw = (request.form.get("client_id") or "").strip()
        branch_id_raw = (request.form.get("branch_id") or "").strip()
        client_id = int(client_id_raw) if client_id_raw.isdigit() else None
        branch_id = int(branch_id_raw) if branch_id_raw.isdigit() else None

        if branch_id and not client_id:
            flash("Please select Client first before selecting Branch.", "danger")
            return render_template(
                "leads/form.html",
                mode="edit", lead=lead,
                statuses=statuses, sources=sources, clients=clients, industries=industries, services=services
            )

        if client_id and branch_id:
            br = ClientBranch.query.get(branch_id)
            if not br or br.client_id != client_id:
                flash("Selected branch does not belong to selected client.", "danger")
                return render_template(
                    "leads/form.html",
                    mode="edit", lead=lead,
                    statuses=statuses, sources=sources, clients=clients, industries=industries, services=services
                )

        lead.client_id = client_id
        lead.branch_id = branch_id
        lead.estimated_closure_date = _parse_date(request.form.get("estimated_closure_date"))

        if not lead.name:
            flash("Lead name is required.", "danger")
            return render_template(
                "leads/form.html",
                mode="edit", lead=lead,
                statuses=statuses, sources=sources, clients=clients, industries=industries, services=services
            )

        db.session.commit()

        if old_name != lead.name:
            log_audit("Lead", lead.id, "UPDATE", "name", old_name, lead.name)

        flash("Lead updated ✅", "success")
        return redirect(url_for("leads.view_lead", lead_id=lead.id))

    return render_template(
        "leads/form.html",
        mode="edit", lead=lead,
        statuses=statuses, sources=sources, clients=clients, industries=industries, services=services
    )


@leads_bp.route("/<int:lead_id>")
@login_required
@require_perm("leads.view")
def view_lead(lead_id):
    lead = Lead.query.get_or_404(lead_id)
    _enforce_lead_access(lead)

    activity_types = ActivityType.query.filter_by(is_active=True).order_by(ActivityType.sort_order.asc()).all()
    activities = (LeadActivity.query
                  .filter_by(lead_id=lead.id)
                  .order_by(LeadActivity.activity_at.desc(), LeadActivity.id.desc())
                  .limit(50).all())
    owner_options = _owner_options_for_current_user()

    return render_template(
        "leads/view.html",
        lead=lead,
        activity_types=activity_types,
        activities=activities,
        owner_options=owner_options
    )


@leads_bp.route("/<int:lead_id>/activities/new", methods=["POST"])
@login_required
@require_perm("activities.create")
def add_activity(lead_id):
    lead = Lead.query.get_or_404(lead_id)
    _enforce_lead_access(lead)

    activity_type_id = request.form.get("activity_type_id")
    subject = (request.form.get("subject") or "").strip()
    outcome = (request.form.get("outcome") or "").strip()
    notes = (request.form.get("notes") or "").strip()

    activity_at_raw = request.form.get("activity_at")
    next_follow_raw = request.form.get("next_follow_up_at")

    def parse_dt(v):
        if not v:
            return None
        return datetime.strptime(v, "%Y-%m-%dT%H:%M")

    activity_at = parse_dt(activity_at_raw) or datetime.utcnow()
    next_follow = parse_dt(next_follow_raw)

    act = LeadActivity(
        lead_id=lead.id,
        activity_type_id=int(activity_type_id) if (activity_type_id or "").isdigit() else None,
        subject=subject,
        outcome=outcome,
        notes=notes,
        activity_at=activity_at,
        next_follow_up_at=next_follow,
        created_by_id=current_user.id
    )
    db.session.add(act)

    new_status_id = request.form.get("new_status_id")
    if (new_status_id or "").isdigit():
        lead.status_id = int(new_status_id)

    db.session.commit()
    flash("Activity added ✅", "success")
    return redirect(url_for("leads.view_lead", lead_id=lead.id))


@leads_bp.route("/follow-ups")
@login_required
@require_perm("activities.view")
def followups():
    now = datetime.utcnow()
    days = int(request.args.get("days", 7))
    end = now + timedelta(days=days)

    allowed_owner_ids = _allowed_lead_owner_ids()

    query = (LeadActivity.query
             .join(Lead, LeadActivity.lead_id == Lead.id)
             .filter(LeadActivity.next_follow_up_at.isnot(None))
             .filter(LeadActivity.next_follow_up_at <= end))

    if allowed_owner_ids is not None:
        query = query.filter(Lead.owner_id.in_(allowed_owner_ids))

    items = (query
             .order_by(LeadActivity.next_follow_up_at.asc())
             .limit(200).all())

    return render_template("leads/followups.html", items=items, days=days)


@leads_bp.route("/<int:lead_id>/assign", methods=["POST"])
@login_required
@require_perm("leads.assign")
def assign_lead(lead_id):
    lead = Lead.query.get_or_404(lead_id)
    _enforce_lead_access(lead)

    new_owner_id = request.form.get("owner_id")
    if not new_owner_id or not new_owner_id.isdigit():
        flash("Invalid user selected.", "danger")
        return redirect(url_for("leads.view_lead", lead_id=lead.id))

    new_owner_id = int(new_owner_id)
    allowed = _allowed_lead_owner_ids()

    if allowed is not None and new_owner_id not in allowed:
        flash("You can only assign leads within your team.", "danger")
        return redirect(url_for("leads.view_lead", lead_id=lead.id))

    old_owner = lead.owner.name if lead.owner else None
    lead.owner_id = new_owner_id
    db.session.commit()

    new_owner = User.query.get(new_owner_id)
    log_audit("Lead", lead.id, "ASSIGN", "owner", old_owner, new_owner.name if new_owner else str(new_owner_id))
    flash("Lead reassigned successfully ✅", "success")
    return redirect(url_for("leads.view_lead", lead_id=lead.id))


# -------- API: Branch dropdown for a client --------
@leads_bp.route("/api/client/<int:client_id>/branches")
@login_required
@require_perm("leads.view")
def api_client_branches(client_id):
    branches = (ClientBranch.query
                .filter_by(client_id=client_id, is_active=True)
                .order_by(ClientBranch.branch_location.asc())
                .all())
    return jsonify([{"id": b.id, "name": b.branch_location} for b in branches])


@leads_bp.route("/import", methods=["GET", "POST"])
@login_required
@require_perm("leads.create")
def import_leads():
    from openpyxl import load_workbook

    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            flash("Please select an Excel file.", "danger")
            return redirect(url_for("leads.import_leads"))

        filename = secure_filename(f.filename)
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            flash("Only .xlsx/.xlsm supported.", "danger")
            return redirect(url_for("leads.import_leads"))

        wb = load_workbook(f, data_only=True)
        ws = wb.active

        # Expected headers (row 1)
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = (ws.cell(row=1, column=c).value or "")
            key = str(v).strip().lower()
            if key:
                headers[key] = c

        def col(key):
            return headers.get(key)

        required = ["name"]
        missing = [k for k in required if col(k) is None]
        if missing:
            flash(f"Missing required columns: {', '.join(missing)}", "danger")
            return redirect(url_for("leads.import_leads"))

        status_map = {s.name.strip().lower(): s.id for s in LeadStatus.query.filter_by(is_active=True).all()}
        source_map = {s.name.strip().lower(): s.id for s in LeadSource.query.filter_by(is_active=True).all()}
        industry_map = {i.name.strip().lower(): i.id for i in Industry.query.filter_by(is_active=True).all()}
        service_map = {s.name.strip().lower(): s.id for s in LeadService.query.filter_by(is_active=True).all()}

        created = 0
        skipped = 0
        errors = []

        next_id = (db.session.query(func.coalesce(func.max(Lead.id), 0)).scalar() or 0) + 1

        for r in range(2, ws.max_row + 1):
            
            try:
                name = (ws.cell(r, col("name")).value or "").strip() if col("name") else ""
                if not name:
                    skipped += 1
                    continue

                company = ws.cell(r, col("company")).value if col("company") else ""
                email = ws.cell(r, col("email")).value if col("email") else ""
                phone_country = ws.cell(r, col("phone_country")).value if col("phone_country") else "+91"
                phone = ws.cell(r, col("phone")).value if col("phone") else ""
                location = ws.cell(r, col("location")).value if col("location") else ""
                website = ws.cell(r, col("website")).value if col("website") else ""
                notes = ws.cell(r, col("notes")).value if col("notes") else ""
                service_name = (ws.cell(r, col("service")).value or "").strip() if col("service") else ""
                service_id = None
                if service_name:
                    key = service_name.lower()
                    service_id = service_map.get(key)
                    if not service_id:
                        new_s = LeadService(name=service_name, is_active=True, sort_order=0)
                        db.session.add(new_s)
                        db.session.flush()
                        service_id = new_s.id
                        service_map[key] = service_id
                industry_name = (ws.cell(r, col("industry")).value or "").strip() if col("industry") else ""
                industry_id = None
                if industry_name:
                    key = industry_name.lower()
                    industry_id = industry_map.get(key)
                    if not industry_id:
                        new_i = Industry(name=industry_name, is_active=True, sort_order=0)
                        db.session.add(new_i)
                        db.session.flush()
                        industry_id = new_i.id
                        industry_map[key] = industry_id
                        

                status_name = (ws.cell(r, col("status")).value or "")
                source_name = (ws.cell(r, col("source")).value or "")
                status_key = str(status_name).strip().lower() if status_name else ""
                source_key = str(source_name).strip().lower() if source_name else ""

                lead_code = f"LD-{next_id:06d}"
                next_id += 1

                lead = Lead(
                    lead_code=lead_code,                    
                    name=str(name).strip(),
                    company=str(company or "").strip(),
                    email=str(email or "").strip().lower(),
                    phone_country=str(phone_country or "+91").strip(),
                    phone=str(phone or "").strip(),
                    location=str(location or "").strip(),
                    industry_id=industry_id,
                    website=str(website or "").strip(),
                    notes=str(notes or "").strip(),
                    owner_id=current_user.id,
                    status_id=status_map.get(status_key) if status_key else None,
                    source_id=source_map.get(source_key) if source_key else None,
                    service_id=service_id,
                )
                db.session.add(lead)
                created += 1
            except Exception as e:
                errors.append(f"Row {r}: {str(e)}")

        db.session.commit()

        msg = f"Imported: {created}, Skipped: {skipped}"
        if errors:
            msg += f" (Errors: {len(errors)})"
        flash(msg, "success" if created else "warning")

        return render_template("leads/import.html", errors=errors)

    return render_template("leads/import.html", errors=None)